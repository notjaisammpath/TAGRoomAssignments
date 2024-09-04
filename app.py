import sys
import sensitive_info
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QComboBox,
    QVBoxLayout,
    QGridLayout,
    QWidget,
    QPushButton,
    QFileDialog,
    QLineEdit,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QMessageBox,
    QStackedWidget,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QPalette, QColor
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

data_by_person = {}
emails_by_person = {}
exam_types = ["Please Select", "AP", "PSAT", "SAT"]
exam = None


class UploadPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout()

        self.title = QLabel("Exam Room Assignments")
        self.title.setFont(QFont("Segoe UI", 24, QFont.Bold))
        self.title.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.title)

        self.instructions = QLabel(
            "Please double check that both files are formatted correctly before uploading. \n If files aren't formated correctly, the program will crash instantly."
        )
        self.instructions.setFont(QFont("Segoe UI", 12))
        self.instructions.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.instructions)

        # Add dropdown for exam type
        self.exam_type_label = QLabel("Select Exam Type:")
        self.exam_type_label.setFont(QFont("Segoe UI", 12))
        self.layout.addWidget(self.exam_type_label)

        self.exam_type_dropdown = QComboBox()
        self.exam_type_dropdown.addItems(exam_types)
        self.exam_type_dropdown.setStyleSheet(
            """
            QComboBox {
                padding: 10px;
                font-size: 16px;
                background-color: #2D2D30;
                color: #FFFFFF;
                border: 1px solid #555555;
                border-radius: 5px;
            }
            QComboBox:hover {
                background-color: #3C3C3C;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                border: none;
                width: 14px;
                height: 14px;
            }
            QComboBox QAbstractItemView {
                background-color: #2D2D30;
                border: 1px solid #555555;
                color: #FFFFFF;
                selection-background-color: #4CAF50;  /* Green accent color */
                selection-color: #FFFFFF;
            }
        """
        )
        self.layout.addWidget(self.exam_type_dropdown)
        

        self.upload_exam_button = QPushButton("Upload Exam Roster Spreadsheet")
        self.upload_exam_button.clicked.connect(self.upload_exam_file)
        self.upload_exam_button.setStyleSheet(
            """
            QPushButton {
                padding: 15px;
                font-size: 16px;
                background-color: #E1E1E1;
                color: #000;
                border: 1px solid #CCC;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #D1D1D1;
            }
        """
        )
        self.layout.addWidget(self.upload_exam_button)

        self.upload_email_button = QPushButton("Upload Student to Email Spreadsheet")
        self.upload_email_button.clicked.connect(self.upload_email_file)
        self.upload_email_button.setStyleSheet(
            """
            QPushButton {
                padding: 15px;
                font-size: 16px;
                background-color: #E1E1E1;
                color: #000;
                border: 1px solid #CCC;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #D1D1D1;
            }
        """
        )
        self.layout.addWidget(self.upload_email_button)

        self.layout.addStretch()
        self.layout.setContentsMargins(50, 50, 50, 50)
        self.layout.setSpacing(20)

        self.setLayout(self.layout)
        self.email_file_uploaded = False
        self.exam_file_uploaded = False

    def get_selected_exam_type(self):
        return self.exam_type_dropdown.currentText()
    
    def upload_exam_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Exam Schedule File",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )
        if file_path:
            self.parse_exam_file(file_path)
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Success")
            msg_box.setText("Exam schedule file uploaded and parsed successfully!")
            msg_box.setStyleSheet("QLabel { color: black; }")
            msg_box.exec_()
            self.exam_file_uploaded = True
            self.check_files_and_navigate()

    def upload_email_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Email File",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )
        if file_path:
            self.parse_email_file(file_path)
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Success")
            msg_box.setText("Email file uploaded and parsed successfully!")
            msg_box.setStyleSheet("QLabel { color: black; }")
            msg_box.exec_()
            self.email_file_uploaded = True
            self.check_files_and_navigate()

    def check_files_and_navigate(self):
        if self.email_file_uploaded and self.exam_file_uploaded and self.get_selected_exam_type() != "Please Select":
            self.parent().setCurrentIndex(1)
            self.parent().parent().set_selected_exam_type(self.get_selected_exam_type())

    def parse_exam_file(self, file_path):
        global data_by_person
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        data_by_person = {}
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                continue
            if row[0] is None:
                continue

            number = row[0]
            last_name = row[1].capitalize()
            first_name = row[2].capitalize()
            exam = row[3]
            date = row[4]
            time = row[5]
            room_number = row[6]
            proctor = row[7]

            full_name = f"{first_name} {last_name}"
            exam_details = {
                "exam": exam,
                "date": date,
                "number": number,
                "am_pm": time,
                "room_number": room_number,
                "proctor": proctor,
            }

            if full_name not in data_by_person:
                data_by_person[full_name] = []
            data_by_person[full_name].append(exam_details)

    def parse_email_file(self, file_path):
        global emails_by_person
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        emails_by_person = {}
        for row in sheet.iter_rows(values_only=True):
            if row[0] is None:
                continue

            last_name = row[0].capitalize().strip()
            first_name = row[1].capitalize().strip()
            email = row[2]

            full_name = f"{first_name} {last_name}"
            emails_by_person[full_name] = email


class SearchPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout()

        self.search_label = QLabel("Search for a Student:")
        self.search_label.setFont(QFont("Segoe UI", 14))
        self.layout.addWidget(self.search_label)

        self.search_input = QLineEdit()
        self.search_input.setFont(QFont("Segoe UI", 12))
        self.layout.addWidget(self.search_input)

        self.search_button = QPushButton("Search")
        self.search_button.clicked.connect(self.search_student)
        self.search_button.setStyleSheet(
            """
            QPushButton {
                padding: 15px;
                font-size: 16px;
                background-color: #E1E1E1;
                color: #000;
                border: 1px solid #CCC;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #D1D1D1;
            }
        """
        )
        self.layout.addWidget(self.search_button)

        self.results_table = QTableWidget()
        self.results_table.setColumnCount(6)
        self.results_table.setHorizontalHeaderLabels(
            ["Exam", "Date", "Time", "Number", "Room Number", "Proctor"]
        )
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.layout.addWidget(self.results_table)

        self.email_button = QPushButton("Send All Emails")
        self.email_button.clicked.connect(self.send_emails)
        self.email_button.setStyleSheet(
            """
            QPushButton {
                padding: 15px;
                font-size: 16px;
                background-color: #E1E1E1;
                color: #000;
                border: 1px solid #CCC;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #D1D1D1;
            }
        """
        )
        self.layout.addWidget(self.email_button)

        self.setLayout(self.layout)

    def search_student(self):
        full_name = " ".join(
            part.capitalize() for part in self.search_input.text().strip().split()
        )
        results = data_by_person.get(full_name, "No data found for this person")
        self.display_results(results, full_name)

    def display_results(self, results, name):
        if results == "No data found for this person":
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("No Data")
            msg_box.setText(f"No data found for {name}")
            msg_box.setStyleSheet("QLabel { color: black; }")
            msg_box.exec_()
            self.results_table.setRowCount(0)
        else:
            self.results_table.setRowCount(len(results))
            for row_idx, result in enumerate(results):
                self.results_table.setItem(row_idx, 0, QTableWidgetItem(result["exam"]))
                self.results_table.setItem(
                    row_idx, 1, QTableWidgetItem(result["date"].strftime("%B %d, %Y"))
                )
                self.results_table.setItem(
                    row_idx, 2, QTableWidgetItem(result["am_pm"])
                )
                self.results_table.setItem(
                    row_idx, 3, QTableWidgetItem(str(int(result["number"])))
                )
                self.results_table.setItem(
                    row_idx, 4, QTableWidgetItem(str(int(result["room_number"])))
                )
                self.results_table.setItem(
                    row_idx, 5, QTableWidgetItem(result["proctor"])
                )

    def send_emails(self):
        self.send_batch_emails(data_by_person, emails_by_person, self.parent().parent().get_selected_exam_type())
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Emails Sent")
        msg_box.setText("Emails have been sent successfully!")
        msg_box.setStyleSheet("QLabel { color: black; }")
        msg_box.exec_()

    def send_batch_emails(self, data_by_person, emails_by_person, exam_type):
        sender_email = sensitive_info.test_sender_email
        password = sensitive_info.test_sender_password

        for student, exams in data_by_person.items():
            receiver_email = emails_by_person.get(student)
            if not receiver_email:
                QMessageBox.warning(
                    self, "Email Not Found", f"No email found for {student}"
                )
                continue

            subject = f"Your {exam_type} Exam Schedule - {student}"
            body = f"""
                <!DOCTYPE html>
            <html>
            <head>
            <title>{exam_type} Exam Schedule for {student}</title>
            <style>
            table {{
            width: 100%;
            border-collapse: collapse;
            }}
            th, td {{
            border: 1px solid #dddddd;
            text-align: left;
            padding: 8px;
            }}
            tr:nth-child(even) {{
            background-color:   
            #f2f2f2;
            }}
            th {{
            background-color: #4CAF50;
            color: white;
            }}
            </style>
            </head>
            <body>   

            <h2> {exam_type} Exam Schedule for {student}</h2>
            <table>
            <tr>
                <th>Exam Name</th>
                <th>Date</th>
                <th>Time</th>
                <th>Number</th>
                <th>Room Number</th>
                <th>Proctor</th>
            </tr>
            """
            for exam in exams:
                body += f"""
                <tr>
                    <td>{exam['exam']}</td>
                    <td>{exam['date'].strftime('%B %d, %Y')}</td>
                    <td>{exam['am_pm']}</td>
                    <td>{round(exam['number'])}</td>
                    <td>{round(exam['room_number'])}</td>
                    <td>{exam['proctor']}</td>
                </tr>
                """
            body += """
                </table>
                </body>
                </html>
            """

            # Send the email
            msg = MIMEMultipart()
            msg["From"] = sender_email
            msg["To"] = receiver_email
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "html"))
            smtp_server = "smtp.gmail.com"
            port = 587

            try:
                server = smtplib.SMTP(smtp_server, port)
                server.starttls()  # Secure the connection
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, msg.as_string())
                print("Email sent successfully!")
            except Exception as e:
                QMessageBox.warning(
                    self, f"Failed to send email to {receiver_email}: {e}"
                )


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Exam Room Assignments")
        self.setGeometry(100, 100, 800, 600)

        self.central_widget = QStackedWidget()
        self.setCentralWidget(self.central_widget)

        self.upload_page = UploadPage(self.central_widget)
        self.search_page = SearchPage(self.central_widget)

        self.central_widget.addWidget(self.upload_page)
        self.central_widget.addWidget(self.search_page)

        self.central_widget.setCurrentIndex(0)

        self.setStyleSheet(
            """
    QMainWindow {
        background-color: #2D2D30;
    }
    QLabel {
        color: #FFFFFF;
        font-family: 'Segoe UI';
    }
    QLineEdit {
        padding: 10px;
        font-size: 14px;
        border: 1px solid #555555;
        border-radius: 5px;
        background-color: #3C3C3C;
        color: #FFFFFF;
    }
    QPushButton {
        padding: 15px;
        font-size: 16px;
        background-color: #3C3C3C;
        color: #FFFFFF;
        border: 1px solid #555555;
        border-radius: 5px;
        font-family: 'Segoe UI';
    }
    QPushButton:hover {
        background-color: #4CAF50;  /* Green accent color */
        border: 1px solid #4CAF50;  /* Green accent color */
    }
    QTableWidget {
        background-color: #3C3C3C;
        border: 1px solid #555555;
        border-radius: 5px;
        color: #FFFFFF;
    }
    QHeaderView::section {
        background-color: #3C3C3C;
        padding: 5px;
        border: none;
        font-family: 'Segoe UI';
        color: #FFFFFF;
    }
    QTableWidget QTableCornerButton::section {
        background-color: #3C3C3C;
        border: none;
    }
    QScrollBar:vertical {
        border: none;
        background: #2D2D30;
        width: 12px;
        margin: 0px 0px 0px 0px;
    }
    QScrollBar::handle:vertical {
        background: #555555;
        min-height: 20px;
        border-radius: 5px;
    }
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
        height: 0px;
    }
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
        background: none;
    }
    QTableView {
        background-color: #3C3C3C;
        alternate-background-color: #2D2D30;
        gridline-color: #555555;
        color: #FFFFFF;
    }
    QHeaderView {
        background-color: #3C3C3C;
        color: #FFFFFF;
    }
    QHeaderView::section {
        background-color: #3C3C3C;
        color: #FFFFFF;
        border: 1px solid #555555;
    }
"""
        )
        self.selected_exam_type = None  # Add this line to store the selected exam type

    def set_selected_exam_type(self, exam_type):
        self.selected_exam_type = exam_type

    def get_selected_exam_type(self):
        return self.selected_exam_type


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
